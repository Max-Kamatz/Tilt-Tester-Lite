# tests/test_test_logger.py
import csv
import os
import tempfile
import pytest
from datetime import datetime
from logger.test_logger import TestLogger


@pytest.fixture
def logger(tmp_path):
    lg = TestLogger()
    lg.start(dir=str(tmp_path))
    yield lg
    lg.close()


def test_temp_file_created(tmp_path):
    lg = TestLogger()
    lg.start(dir=str(tmp_path))
    assert lg.temp_path is not None
    assert os.path.exists(lg.temp_path)
    lg.close()


def test_temp_file_deleted_on_close(tmp_path):
    lg = TestLogger()
    lg.start(dir=str(tmp_path))
    path = lg.temp_path
    lg.close()
    assert not os.path.exists(path)


def test_log_writes_row(logger, tmp_path):
    ts = datetime(2026, 3, 30, 12, 0, 0)
    logger.log(ts, "PTZ", "Cycle Complete", "")
    with open(logger.temp_path, newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    # header + 1 data row
    assert len(rows) == 2
    assert rows[1][2] == "Cycle Complete"


def test_log_multiple_rows(logger):
    ts = datetime(2026, 3, 30, 12, 0, 0)
    for i in range(5):
        logger.log(ts, "PTZ", f"Event {i}", "")
    with open(logger.temp_path, newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    assert len(rows) == 6  # header + 5


def test_export_csv(logger, tmp_path):
    ts = datetime(2026, 3, 30, 12, 0, 0)
    logger.log(ts, "PTZ", "Test Start", "")
    out = str(tmp_path / "export.csv")
    logger.export_csv(out)
    with open(out, newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    assert rows[0] == ["Timestamp", "Source", "Event Type", "Detail"]
    assert rows[1][2] == "Test Start"


def test_export_excel(logger, tmp_path):
    from openpyxl import load_workbook
    ts = datetime(2026, 3, 30, 12, 0, 0)
    logger.log(ts, "10.10.10.2", "Ping Loss", "")
    out = str(tmp_path / "export.xlsx")
    logger.export_excel(out)
    wb = load_workbook(out)
    ws = wb.active
    assert ws.cell(1, 1).value == "Timestamp"
    assert ws.cell(2, 3).value == "Ping Loss"
