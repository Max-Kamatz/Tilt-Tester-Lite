# logger/test_logger.py
import csv
import os
import shutil
import tempfile
from datetime import datetime

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

HEADERS = ["Timestamp", "Source", "Event Type", "Detail"]


class TestLogger:
    def __init__(self):
        self.temp_path: str | None = None
        self._fd: int | None = None

    def start(self, dir: str | None = None) -> None:
        self._fd, self.temp_path = tempfile.mkstemp(
            suffix=".tmp.csv", prefix="TiltTesterLite_", dir=dir
        )
        with os.fdopen(self._fd, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(HEADERS)
        self._fd = None

    def log(self, timestamp: datetime, source: str,
            event_type: str, detail: str) -> None:
        ts_str = timestamp.strftime("%Y-%m-%d %H:%M:%S.") + \
                 f"{timestamp.microsecond // 1000:03d}"
        with open(self.temp_path, "a", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow([ts_str, source, event_type, detail])

    def export_csv(self, path: str) -> None:
        shutil.copy2(self.temp_path, path)

    def export_excel(self, path: str) -> None:
        wb = Workbook()
        ws = wb.active
        with open(self.temp_path, newline="", encoding="utf-8") as f:
            for row in csv.reader(f):
                ws.append(row)
        ws.auto_filter.ref = ws.dimensions
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = \
                min(max_len + 2, 60)
        wb.save(path)

    def close(self) -> None:
        if self.temp_path and os.path.exists(self.temp_path):
            os.remove(self.temp_path)
        self.temp_path = None
